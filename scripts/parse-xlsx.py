import openpyxl
import json
import os

xlsx_path = "scratch/source.xlsx"
json_path = "scratch/sheets.json"

if not os.path.exists(xlsx_path):
    print(f"Error: {xlsx_path} not found")
    exit(1)

import datetime

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
sheets_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    
    # Let's clean up empty trailing rows/columns
    max_row = ws.max_row
    max_col = ws.max_column
    
    for r in range(1, max_row + 1):
        row_values = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                val = ""
            elif isinstance(val, (datetime.datetime, datetime.date)):
                val = val.isoformat()
            row_values.append(val)
        rows.append(row_values)
        
    sheets_data[sheet_name] = {
        "rows": rows,
        "nrows": len(rows),
        "ncols": max_col
    }

os.makedirs(os.path.dirname(json_path), exist_ok=True)
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(sheets_data, f, ensure_ascii=False, indent=2)

print("Successfully converted Excel to scratch/sheets.json!")
